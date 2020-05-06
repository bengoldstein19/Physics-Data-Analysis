# download openpyxl from here https://pypi.python.org/pypi/openpyxl
# install by typing at cmd prompt "python C:\Python27\openpyxl-2.4.7\setup.py install"

from openpyxl import Workbook
from visual import sphere,rate,arrow,vector,norm,mag,color
from math import cos,sin,pi,sqrt
from numpy import arange,array,dot

wb = Workbook()
ws = wb.active
ws.title = 'Scrambler Data'

ws['A1'] = "THETA"
ws['B1'] = "VELOCITY"
ws['C1'] = "TANGACCEL"
ws['D1'] = "INWARDACCEL"

data = []

r1 = input("Enter the value of the major radius of the Scrambler: ")
ride_radius = r1
r2 = input("Enter the value of the minor radius of the Scrambler: ")
ride_radius2 = r2
p_temp = input("Enter the value of p, which is the ratio of the period of the minor axis to the major axis. \
    A negative value indicates the circles go in opposite directions: ")
p = p_temp
step_interval = input("Enter the radian measure between each data point (small numbers like 0.1 make smoother plots): ")
T_step = step_interval
s = sphere(make_trail=True, pos=[ride_radius+ride_radius2,0,0],radius=0.1)
s.trail_object.color=color.yellow
a_tangential = arrow(pos=[1,0,0], axis=[0,1,0],color=color.red)
a_inward = arrow(pos=[1,0,0], axis=[-1,0,0], color=color.blue)
for theta in arange(0,100*pi,T_step):
    rate(60)
    r = vector(ride_radius*cos(theta) + ride_radius2*cos(p*theta), \
         ride_radius*sin(theta) + ride_radius2*sin(p*theta), 0)
    velocity = vector(-ride_radius*sin(theta)-ride_radius2*p*sin(p*theta), \
         ride_radius*cos(theta)+ride_radius2*p*cos(p*theta),0)
    vel = str(velocity).strip('[]')
    acceleration = vector(-ride_radius*cos(theta) - ride_radius2*p**2*cos(p*theta), \
         -ride_radius*sin(theta) - ride_radius2*p**2*sin(p*theta), 0)
    tangential_acceleration = (dot(velocity, acceleration)/mag(velocity))*norm(velocity)
    tang_accel = str(tangential_acceleration).strip('[]')
    inward_acceleration = acceleration - tangential_acceleration
    in_accel = str(inward_acceleration).strip('[]')
    s.pos = r
    a_tangential.pos = r
    a_inward.pos = r
    a_tangential.axis = tangential_acceleration
    a_inward.axis = inward_acceleration
    data.append([theta, vel, tang_accel, in_accel])


for row in data:
    ws.append(row)

ws.append (["Major Radius: ", r1])
ws.append (["Minor Radius: ", r2])
ws.append (["Value of p: ", p_temp])
ws.append (["Step Interval: ", step_interval])

# Save the file
wb.save("C:\Users\Sarah\Documents\Hopkins\Python\Phil\sample.xlsx")
